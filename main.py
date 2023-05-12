import openpyxl

# Open the text file for reading
with open('input.txt', 'r') as file:
    input_lines = file.readlines()

# Create a new workbook and sheet
workbook = openpyxl.Workbook()
sheet = workbook.active

xlsx_row = 0
column_letter_count = 25
product_list = []
product_lines_group = []
temp_line_group = []

# split by product
for i, line in enumerate(input_lines):
    if i < 10:
        continue
    if line.startswith('Productgroep'):
        if len(temp_line_group) != 0:
            product_lines_group.append(temp_line_group)
        temp_line_group = []
    temp_line_group.append(line)
product_lines_group.append(temp_line_group)

# write head line of the xlsx
sheet.cell(row=xlsx_row + 1, column=1).value = "Productgroep"
sheet.cell(row=xlsx_row + 1, column=2).value = "Maximumprijs"
xlsx_row = xlsx_row + 1

for i, product_lines in enumerate(product_lines_group):
    write_product = True
    find_product_name = False
    product_name = ""

    for j, product_line in enumerate(product_lines):        
        
        first_column = product_line[: column_letter_count].strip()

        if first_column == "Productgroep":
            continue
        elif first_column == "Registratienummer":
            find_product_name = True
            continue
        
        if find_product_name:
            line_register_number = product_line[: column_letter_count].strip()
            line_article_name = product_line[column_letter_count:].strip()

            if write_product == True:
                article_name = line_article_name
                register_number = line_register_number
            else:
                article_name = article_name + line_article_name

            if j == len(product_lines) - 1:
                write_product = True
            else:
                next_product_line = product_lines[j + 1]
                next_register_number = next_product_line[: column_letter_count].strip()
                write_product = next_register_number != ""

            if write_product:
                sheet.cell(row=xlsx_row + 1, column=1).value = product_name
                sheet.cell(row=xlsx_row + 1, column=2).value = maxium_price
                sheet.cell(row=xlsx_row + 1, column=3).value = unit
                sheet.cell(row=xlsx_row + 1, column=4).value = register_number
                sheet.cell(row=xlsx_row + 1, column=5).value = article_name
                xlsx_row = xlsx_row + 1
        else:
            if product_name == "":
                product_name = product_line[:column_letter_count * 2].strip()
                maxium_price = product_line[column_letter_count * 2:].strip()
                maxium_price_spread = maxium_price.split(" ")
                maxium_price = maxium_price_spread[0]
                maxium_price_spread.pop(0)
                unit = "".join(maxium_price_spread)
            else :
                product_name = product_name + " " + product_line[: column_letter_count].strip()

# Save the workbook as an xlsx file
workbook.save('output.xlsx')